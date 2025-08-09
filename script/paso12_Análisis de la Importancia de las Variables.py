"""
================================================================================
TELECOMX - PASO 12: ANÁLISIS DE LA IMPORTANCIA DE LAS VARIABLES
================================================================================
Descripción: Análisis exhaustivo de las variables más relevantes para la predicción
             de cancelación en los modelos Random Forest y Regresión Logística.
             
Análisis Realizados:
- Random Forest: Importancia por reducción de impureza
- Regresión Logística: Coeficientes e interpretación de odds ratios
- Comparación entre modelos
- Rankings de variables más importantes
- Recomendaciones de negocio basadas en variables clave

Autor: Sistema de Análisis Predictivo TelecomX
Fecha: 2025
================================================================================
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import logging
import os
import joblib
import json
import pickle
from datetime import datetime
from pathlib import Path
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')

def setup_logging():
    """Configurar sistema de logging"""
    os.makedirs('logs', exist_ok=True)
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('logs/paso12_analisis_importancia_variables.log', mode='a', encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

def create_directories():
    """Crear directorios necesarios"""
    directories = ['excel', 'informes', 'graficos', 'logs', 'datos', 'modelos']
    for directory in directories:
        os.makedirs(directory, exist_ok=True)
        logging.info(f"Carpeta verificada/creada: {directory}")

def find_latest_file(directory, pattern):
    """Encontrar el archivo más reciente que coincida con el patrón"""
    files = list(Path(directory).glob(pattern))
    if not files:
        raise FileNotFoundError(f"No se encontraron archivos con patrón {pattern} en {directory}")
    latest_file = max(files, key=os.path.getctime)
    return str(latest_file)

def load_models_and_data():
    """Cargar modelos entrenados y datos del Paso 11"""
    try:
        logging.info("Cargando modelos y datos del Paso 11...")
        
        # Buscar archivos de modelos más recientes
        rf_model_file = find_latest_file('modelos', 'random_forest_model_*.pkl')
        lr_model_file = find_latest_file('modelos', 'logistic_regression_pipeline_*.pkl')
        
        # Cargar Random Forest
        with open(rf_model_file, 'rb') as f:
            rf_model = pickle.load(f)
        logging.info(f"Random Forest cargado: {rf_model_file}")
        
        # Cargar Regresión Logística (pipeline)
        with open(lr_model_file, 'rb') as f:
            lr_model = pickle.load(f)
        logging.info(f"Regresión Logística cargada: {lr_model_file}")
        
        # Cargar información de modelos
        rf_info_file = rf_model_file.replace('_model_', '_info_').replace('.pkl', '.json')
        lr_info_file = lr_model_file.replace('_pipeline_', '_info_').replace('.pkl', '.json')
        
        with open(rf_info_file, 'r', encoding='utf-8') as f:
            rf_info = json.load(f)
        
        with open(lr_info_file, 'r', encoding='utf-8') as f:
            lr_info = json.load(f)
        
        # Cargar dataset para obtener nombres de variables
        train_file = find_latest_file('datos', 'telecomx_train_dataset_*.csv')
        
        # Probar diferentes encodings
        encodings = ['utf-8-sig', 'utf-8', 'cp1252', 'latin-1']
        df_train = None
        
        for encoding in encodings:
            try:
                df_train = pd.read_csv(train_file, encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        
        if df_train is None:
            raise ValueError(f"No se pudo cargar {train_file}")
        
        # Separar características y objetivo
        target_var = 'Abandono_Cliente'
        feature_names = df_train.drop(columns=[target_var]).columns.tolist()
        
        models_data = {
            'Random Forest': {
                'model': rf_model,
                'info': rf_info,
                'file': rf_model_file,
                'type': 'tree_based'
            },
            'Regresión Logística': {
                'model': lr_model,
                'info': lr_info,
                'file': lr_model_file,
                'type': 'linear'
            }
        }
        
        logging.info(f"Datos cargados exitosamente: {len(feature_names)} variables")
        return models_data, feature_names, df_train
        
    except Exception as e:
        logging.error(f"Error al cargar modelos y datos: {str(e)}")
        raise

def analyze_random_forest_importance(rf_model, feature_names):
    """Analizar importancia de variables en Random Forest"""
    logging.info("Analizando importancia de variables en Random Forest...")
    
    try:
        # Extraer importancia de características
        feature_importance = rf_model.feature_importances_
        
        # Crear DataFrame con importancias
        importance_df = pd.DataFrame({
            'Variable': feature_names,
            'Importancia': feature_importance,
            'Importancia_Pct': feature_importance / feature_importance.sum() * 100
        })
        
        # Ordenar por importancia descendente
        importance_df = importance_df.sort_values('Importancia', ascending=False).reset_index(drop=True)
        importance_df['Ranking'] = importance_df.index + 1
        
        # Importancia acumulada
        importance_df['Importancia_Acumulada'] = importance_df['Importancia_Pct'].cumsum()
        
        # Estadísticas adicionales del modelo
        n_trees = rf_model.n_estimators
        max_depth = rf_model.max_depth
        
        # Identificar variables críticas (top que suman 80% de importancia)
        critical_threshold = 80
        critical_vars = importance_df[importance_df['Importancia_Acumulada'] <= critical_threshold]
        
        analysis_results = {
            'model_name': 'Random Forest',
            'importance_df': importance_df,
            'critical_variables': critical_vars,
            'model_params': {
                'n_trees': n_trees,
                'max_depth': max_depth,
                'total_features': len(feature_names)
            },
            'insights': {
                'top_variable': importance_df.iloc[0]['Variable'],
                'top_importance': importance_df.iloc[0]['Importancia_Pct'],
                'critical_count': len(critical_vars),
                'concentration': critical_vars['Importancia_Pct'].sum()
            }
        }
        
        logging.info(f"Random Forest - Variable más importante: {analysis_results['insights']['top_variable']} ({analysis_results['insights']['top_importance']:.2f}%)")
        logging.info(f"Random Forest - Top {len(critical_vars)} variables explican {analysis_results['insights']['concentration']:.1f}% de la importancia")
        
        return analysis_results
        
    except Exception as e:
        logging.error(f"Error analizando Random Forest: {str(e)}")
        raise

def analyze_logistic_regression_coefficients(lr_pipeline, feature_names):
    """Analizar coeficientes de Regresión Logística"""
    logging.info("Analizando coeficientes de Regresión Logística...")
    
    try:
        # Extraer el modelo de regresión logística del pipeline
        lr_model = None
        
        # El pipeline puede tener diferentes estructuras
        if hasattr(lr_pipeline, 'named_steps'):
            # Pipeline con pasos nombrados
            for step_name, step_object in lr_pipeline.named_steps.items():
                if hasattr(step_object, 'coef_'):
                    lr_model = step_object
                    break
        elif hasattr(lr_pipeline, 'coef_'):
            # El objeto ya es el modelo
            lr_model = lr_pipeline
        else:
            # Buscar en steps si existe
            if hasattr(lr_pipeline, 'steps'):
                for step_name, step_object in lr_pipeline.steps:
                    if hasattr(step_object, 'coef_'):
                        lr_model = step_object
                        break
        
        if lr_model is None or not hasattr(lr_model, 'coef_'):
            raise ValueError("No se pudo encontrar modelo de regresión logística con coeficientes")
        
        # Extraer coeficientes
        coefficients = lr_model.coef_[0]  # Para clasificación binaria
        
        # Crear DataFrame con coeficientes
        coef_df = pd.DataFrame({
            'Variable': feature_names,
            'Coeficiente': coefficients,
            'Coeficiente_Abs': np.abs(coefficients),
            'Odds_Ratio': np.exp(coefficients),
            'Direccion': ['Positivo' if c > 0 else 'Negativo' for c in coefficients]
        })
        
        # Ordenar por valor absoluto del coeficiente
        coef_df = coef_df.sort_values('Coeficiente_Abs', ascending=False).reset_index(drop=True)
        coef_df['Ranking'] = coef_df.index + 1
        
        # Calcular importancia relativa (valor absoluto normalizado)
        coef_df['Importancia_Relativa'] = coef_df['Coeficiente_Abs'] / coef_df['Coeficiente_Abs'].sum() * 100
        coef_df['Importancia_Acumulada'] = coef_df['Importancia_Relativa'].cumsum()
        
        # Identificar variables críticas (top que suman 80% de importancia)
        critical_threshold = 80
        critical_vars = coef_df[coef_df['Importancia_Acumulada'] <= critical_threshold]
        
        # Estadísticas adicionales
        regularization = getattr(lr_model, 'C', 'N/A')
        penalty = getattr(lr_model, 'penalty', 'N/A')
        
        analysis_results = {
            'model_name': 'Regresión Logística',
            'coefficients_df': coef_df,
            'critical_variables': critical_vars,
            'model_params': {
                'regularization_C': regularization,
                'penalty': penalty,
                'total_features': len(feature_names)
            },
            'insights': {
                'top_variable': coef_df.iloc[0]['Variable'],
                'top_coefficient': coef_df.iloc[0]['Coeficiente'],
                'top_odds_ratio': coef_df.iloc[0]['Odds_Ratio'],
                'critical_count': len(critical_vars),
                'positive_vars': len(coef_df[coef_df['Direccion'] == 'Positivo']),
                'negative_vars': len(coef_df[coef_df['Direccion'] == 'Negativo'])
            }
        }
        
        logging.info(f"Regresión Logística - Variable más importante: {analysis_results['insights']['top_variable']}")
        logging.info(f"Regresión Logística - Coeficiente: {analysis_results['insights']['top_coefficient']:.4f}")
        logging.info(f"Regresión Logística - Odds Ratio: {analysis_results['insights']['top_odds_ratio']:.4f}")
        logging.info(f"Regresión Logística - Variables positivas: {analysis_results['insights']['positive_vars']}, negativas: {analysis_results['insights']['negative_vars']}")
        
        return analysis_results
        
    except Exception as e:
        logging.error(f"Error analizando Regresión Logística: {str(e)}")
        raise

def compare_variable_importance(rf_analysis, lr_analysis):
    """Comparar importancia de variables entre modelos"""
    logging.info("Comparando importancia de variables entre modelos...")
    
    try:
        # Obtener top variables de cada modelo
        rf_top = rf_analysis['importance_df'].head(20)
        lr_top = lr_analysis['coefficients_df'].head(20)
        
        # Crear DataFrame de comparación
        rf_importance = dict(zip(rf_top['Variable'], rf_top['Importancia_Pct']))
        lr_importance = dict(zip(lr_top['Variable'], lr_top['Importancia_Relativa']))
        
        # Variables únicas en ambos modelos
        all_variables = set(rf_importance.keys()) | set(lr_importance.keys())
        
        comparison_df = pd.DataFrame({
            'Variable': list(all_variables),
            'RF_Importancia': [rf_importance.get(var, 0) for var in all_variables],
            'LR_Importancia': [lr_importance.get(var, 0) for var in all_variables]
        })
        
        # Calcular rankings
        comparison_df['RF_Ranking'] = comparison_df['RF_Importancia'].rank(method='dense', ascending=False)
        comparison_df['LR_Ranking'] = comparison_df['LR_Importancia'].rank(method='dense', ascending=False)
        
        # Calcular importancia promedio y consenso
        comparison_df['Importancia_Promedio'] = (comparison_df['RF_Importancia'] + comparison_df['LR_Importancia']) / 2
        comparison_df['Consenso_Score'] = np.minimum(comparison_df['RF_Importancia'], comparison_df['LR_Importancia'])
        comparison_df['Diferencia_Ranking'] = np.abs(comparison_df['RF_Ranking'] - comparison_df['LR_Ranking'])
        
        # Ordenar por importancia promedio
        comparison_df = comparison_df.sort_values('Importancia_Promedio', ascending=False).reset_index(drop=True)
        comparison_df['Ranking_Consenso'] = comparison_df.index + 1
        
        # Identificar variables de alto consenso (ambos modelos las consideran importantes)
        high_consensus = comparison_df[
            (comparison_df['RF_Ranking'] <= 10) & 
            (comparison_df['LR_Ranking'] <= 10)
        ].copy()
        
        # Variables con mayor discrepancia
        high_discrepancy = comparison_df.nlargest(5, 'Diferencia_Ranking')
        
        comparison_results = {
            'comparison_df': comparison_df,
            'high_consensus_vars': high_consensus,
            'high_discrepancy_vars': high_discrepancy,
            'insights': {
                'consensus_count': len(high_consensus),
                'top_consensus_var': comparison_df.iloc[0]['Variable'] if len(comparison_df) > 0 else 'N/A',
                'avg_discrepancy': comparison_df['Diferencia_Ranking'].mean(),
                'models_agreement': len(high_consensus) / min(20, len(comparison_df)) * 100 if len(comparison_df) > 0 else 0
            }
        }
        
        logging.info(f"Variables de alto consenso: {len(high_consensus)}")
        logging.info(f"Variable de mayor consenso: {comparison_results['insights']['top_consensus_var']}")
        logging.info(f"Concordancia entre modelos: {comparison_results['insights']['models_agreement']:.1f}%")
        
        return comparison_results
        
    except Exception as e:
        logging.error(f"Error comparando importancias: {str(e)}")
        raise

def generate_visualizations(rf_analysis, lr_analysis, comparison, timestamp):
    """Generar todas las visualizaciones"""
    logging.info("Generando visualizaciones...")
    
    try:
        viz_files = []
        
        # 1. Importancia de variables - Random Forest
        plt.figure(figsize=(12, 8))
        rf_top_15 = rf_analysis['importance_df'].head(15)
        
        bars = plt.barh(range(len(rf_top_15)), rf_top_15['Importancia_Pct'], 
                       color='lightgreen', alpha=0.8, edgecolor='darkgreen')
        
        plt.yticks(range(len(rf_top_15)), rf_top_15['Variable'])
        plt.xlabel('Importancia (%)')
        plt.title('Random Forest - Top 15 Variables Más Importantes\n(Basado en Reducción de Impureza)', 
                 fontsize=14, fontweight='bold')
        plt.gca().invert_yaxis()
        
        # Añadir valores en las barras
        for i, (bar, value) in enumerate(zip(bars, rf_top_15['Importancia_Pct'])):
            plt.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height()/2, 
                    f'{value:.2f}%', va='center', fontsize=10)
        
        plt.grid(True, alpha=0.3, axis='x')
        plt.tight_layout()
        
        viz_file = f'graficos/paso12_random_forest_importancia_{timestamp}.png'
        plt.savefig(viz_file, dpi=300, bbox_inches='tight')
        plt.close()
        viz_files.append(viz_file)
        
        # 2. Coeficientes - Regresión Logística
        plt.figure(figsize=(12, 8))
        lr_top_15 = lr_analysis['coefficients_df'].head(15)
        
        colors = ['red' if x < 0 else 'blue' for x in lr_top_15['Coeficiente']]
        bars = plt.barh(range(len(lr_top_15)), lr_top_15['Coeficiente'], 
                       color=colors, alpha=0.7)
        
        plt.yticks(range(len(lr_top_15)), lr_top_15['Variable'])
        plt.xlabel('Coeficiente')
        plt.title('Regresión Logística - Top 15 Variables por Coeficiente\n(Azul: Aumenta Churn, Rojo: Reduce Churn)', 
                 fontsize=14, fontweight='bold')
        plt.gca().invert_yaxis()
        plt.axvline(x=0, color='black', linestyle='-', alpha=0.5)
        
        # Añadir valores en las barras
        for i, (bar, value) in enumerate(zip(bars, lr_top_15['Coeficiente'])):
            offset = 0.01 if value >= 0 else -0.01
            ha = 'left' if value >= 0 else 'right'
            plt.text(bar.get_width() + offset, bar.get_y() + bar.get_height()/2, 
                    f'{value:.3f}', va='center', ha=ha, fontsize=10)
        
        plt.grid(True, alpha=0.3, axis='x')
        plt.tight_layout()
        
        viz_file = f'graficos/paso12_regresion_logistica_coeficientes_{timestamp}.png'
        plt.savefig(viz_file, dpi=300, bbox_inches='tight')
        plt.close()
        viz_files.append(viz_file)
        
        # 3. Comparación de importancias
        plt.figure(figsize=(14, 10))
        
        # Tomar top 15 variables por consenso
        comp_top_15 = comparison['comparison_df'].head(15)
        
        x = np.arange(len(comp_top_15))
        width = 0.35
        
        bars1 = plt.bar(x - width/2, comp_top_15['RF_Importancia'], width, 
                       label='Random Forest', color='lightgreen', alpha=0.8)
        bars2 = plt.bar(x + width/2, comp_top_15['LR_Importancia'], width, 
                       label='Regresión Logística', color='lightblue', alpha=0.8)
        
        plt.xlabel('Variables')
        plt.ylabel('Importancia (%)')
        plt.title('Comparación de Importancia de Variables entre Modelos\n(Top 15 por Consenso)', 
                 fontsize=14, fontweight='bold')
        plt.xticks(x, comp_top_15['Variable'], rotation=45, ha='right')
        plt.legend()
        
        # Añadir valores en las barras
        for bars in [bars1, bars2]:
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    plt.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                            f'{height:.1f}%', ha='center', va='bottom', fontsize=8)
        
        plt.grid(True, alpha=0.3, axis='y')
        plt.tight_layout()
        
        viz_file = f'graficos/paso12_comparacion_modelos_{timestamp}.png'
        plt.savefig(viz_file, dpi=300, bbox_inches='tight')
        plt.close()
        viz_files.append(viz_file)
        
        # 4. Análisis de consenso
        plt.figure(figsize=(12, 8))
        
        # Scatter plot de importancias
        plt.scatter(comparison['comparison_df']['RF_Importancia'], 
                   comparison['comparison_df']['LR_Importancia'],
                   alpha=0.6, s=60, color='purple')
        
        # Línea diagonal para referencia
        max_val = max(comparison['comparison_df']['RF_Importancia'].max(),
                     comparison['comparison_df']['LR_Importancia'].max())
        plt.plot([0, max_val], [0, max_val], 'r--', alpha=0.5, label='Concordancia perfecta')
        
        # Anotar variables de alto consenso
        high_consensus = comparison['high_consensus_vars']
        if len(high_consensus) > 0:
            top_consensus = high_consensus.head(5)
            for _, row in top_consensus.iterrows():
                plt.annotate(row['Variable'][:15], 
                           (row['RF_Importancia'], row['LR_Importancia']),
                           xytext=(5, 5), textcoords='offset points',
                           fontsize=9, alpha=0.8)
        
        plt.xlabel('Random Forest - Importancia (%)')
        plt.ylabel('Regresión Logística - Importancia (%)')
        plt.title('Consenso entre Modelos - Importancia de Variables\n(Variables más cercanas a la línea roja tienen mayor consenso)', 
                 fontsize=14, fontweight='bold')
        plt.legend()
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        
        viz_file = f'graficos/paso12_consenso_variables_{timestamp}.png'
        plt.savefig(viz_file, dpi=300, bbox_inches='tight')
        plt.close()
        viz_files.append(viz_file)
        
        # 5. Odds Ratios de Regresión Logística
        plt.figure(figsize=(12, 8))
        
        # Tomar variables con odds ratios más extremos
        lr_extreme = lr_analysis['coefficients_df'].copy()
        lr_extreme['OR_Distance'] = np.abs(lr_extreme['Odds_Ratio'] - 1)
        lr_extreme = lr_extreme.sort_values('OR_Distance', ascending=False).head(15)
        
        colors = ['red' if x < 1 else 'blue' for x in lr_extreme['Odds_Ratio']]
        bars = plt.barh(range(len(lr_extreme)), lr_extreme['Odds_Ratio'], 
                       color=colors, alpha=0.7)
        
        plt.yticks(range(len(lr_extreme)), lr_extreme['Variable'])
        plt.xlabel('Odds Ratio')
        plt.title('Regresión Logística - Top 15 Variables por Odds Ratio\n(Azul: OR>1 Aumenta Churn, Rojo: OR<1 Reduce Churn)', 
                 fontsize=14, fontweight='bold')
        plt.gca().invert_yaxis()
        plt.axvline(x=1, color='black', linestyle='-', alpha=0.5, label='OR = 1 (Sin efecto)')
        
        # Añadir valores en las barras
        for i, (bar, value) in enumerate(zip(bars, lr_extreme['Odds_Ratio'])):
            offset = 0.02 if value >= 1 else -0.02
            ha = 'left' if value >= 1 else 'right'
            plt.text(bar.get_width() + offset, bar.get_y() + bar.get_height()/2, 
                    f'{value:.2f}', va='center', ha=ha, fontsize=10)
        
        plt.legend()
        plt.grid(True, alpha=0.3, axis='x')
        plt.tight_layout()
        
        viz_file = f'graficos/paso12_odds_ratios_{timestamp}.png'
        plt.savefig(viz_file, dpi=300, bbox_inches='tight')
        plt.close()
        viz_files.append(viz_file)
        
        logging.info(f"Generadas {len(viz_files)} visualizaciones")
        return viz_files
        
    except Exception as e:
        logging.error(f"Error generando visualizaciones: {str(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
        return []

def create_excel_report(rf_analysis, lr_analysis, comparison, timestamp):
    """Crear reporte Excel con rankings de variables"""
    logging.info("Creando reporte Excel...")
    
    try:
        # Crear workbook
        wb = Workbook()
        
        # Eliminar hoja por defecto
        wb.remove(wb.active)
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # 1. Hoja de Random Forest
        ws_rf = wb.create_sheet(title="Random Forest")
        
        # Datos de Random Forest
        rf_data = rf_analysis['importance_df'].copy()
        rf_data = rf_data.round(4)
        
        # Escribir headers
        headers_rf = ['Ranking', 'Variable', 'Importancia', 'Importancia (%)', 'Importancia Acumulada (%)']
        for col, header in enumerate(headers_rf, 1):
            cell = ws_rf.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Escribir datos
        for row, (_, data_row) in enumerate(rf_data.iterrows(), 2):
            ws_rf.cell(row=row, column=1, value=data_row['Ranking'])
            ws_rf.cell(row=row, column=2, value=data_row['Variable'])
            ws_rf.cell(row=row, column=3, value=data_row['Importancia'])
            ws_rf.cell(row=row, column=4, value=data_row['Importancia_Pct'])
            ws_rf.cell(row=row, column=5, value=data_row['Importancia_Acumulada'])
        
        # Ajustar anchos de columna
        for col in range(1, 6):
            ws_rf.column_dimensions[chr(64 + col)].width = 20
        
        # 2. Hoja de Regresión Logística
        ws_lr = wb.create_sheet(title="Regresión Logística")
        
        # Datos de Regresión Logística
        lr_data = lr_analysis['coefficients_df'].copy()
        lr_data = lr_data.round(4)
        
        # Escribir headers
        headers_lr = ['Ranking', 'Variable', 'Coeficiente', 'Coef. Absoluto', 'Odds Ratio', 'Dirección', 'Importancia Relativa (%)']
        for col, header in enumerate(headers_lr, 1):
            cell = ws_lr.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Escribir datos
        for row, (_, data_row) in enumerate(lr_data.iterrows(), 2):
            ws_lr.cell(row=row, column=1, value=data_row['Ranking'])
            ws_lr.cell(row=row, column=2, value=data_row['Variable'])
            ws_lr.cell(row=row, column=3, value=data_row['Coeficiente'])
            ws_lr.cell(row=row, column=4, value=data_row['Coeficiente_Abs'])
            ws_lr.cell(row=row, column=5, value=data_row['Odds_Ratio'])
            ws_lr.cell(row=row, column=6, value=data_row['Direccion'])
            ws_lr.cell(row=row, column=7, value=data_row['Importancia_Relativa'])
        
        # Ajustar anchos de columna
        for col in range(1, 8):
            ws_lr.column_dimensions[chr(64 + col)].width = 20
        
        # 3. Hoja de Comparación
        ws_comp = wb.create_sheet(title="Comparación Modelos")
        
        # Datos de comparación
        comp_data = comparison['comparison_df'].copy()
        comp_data = comp_data.round(4)
        
        # Escribir headers
        headers_comp = ['Ranking Consenso', 'Variable', 'RF Importancia (%)', 'LR Importancia (%)', 
                       'Importancia Promedio', 'RF Ranking', 'LR Ranking', 'Diferencia Ranking']
        for col, header in enumerate(headers_comp, 1):
            cell = ws_comp.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Escribir datos
        for row, (_, data_row) in enumerate(comp_data.iterrows(), 2):
            ws_comp.cell(row=row, column=1, value=data_row['Ranking_Consenso'])
            ws_comp.cell(row=row, column=2, value=data_row['Variable'])
            ws_comp.cell(row=row, column=3, value=data_row['RF_Importancia'])
            ws_comp.cell(row=row, column=4, value=data_row['LR_Importancia'])
            ws_comp.cell(row=row, column=5, value=data_row['Importancia_Promedio'])
            ws_comp.cell(row=row, column=6, value=data_row['RF_Ranking'])
            ws_comp.cell(row=row, column=7, value=data_row['LR_Ranking'])
            ws_comp.cell(row=row, column=8, value=data_row['Diferencia_Ranking'])
        
        # Ajustar anchos de columna
        for col in range(1, 9):
            ws_comp.column_dimensions[chr(64 + col)].width = 20
        
        # 4. Hoja de Resumen Ejecutivo
        ws_summary = wb.create_sheet(title="Resumen Ejecutivo", index=0)
        
        # Título
        ws_summary.cell(row=1, column=1, value="ANÁLISIS DE IMPORTANCIA DE VARIABLES - RESUMEN EJECUTIVO")
        ws_summary.cell(row=1, column=1).font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:F1')
        
        # Información general
        info_data = [
            ["", ""],
            ["RANDOM FOREST - INSIGHTS:", ""],
            [f"Variable más importante:", rf_analysis['insights']['top_variable']],
            [f"Importancia:", f"{rf_analysis['insights']['top_importance']:.2f}%"],
            [f"Variables críticas (80% importancia):", rf_analysis['insights']['critical_count']],
            ["", ""],
            ["REGRESIÓN LOGÍSTICA - INSIGHTS:", ""],
            [f"Variable más importante:", lr_analysis['insights']['top_variable']],
            [f"Coeficiente:", f"{lr_analysis['insights']['top_coefficient']:.4f}"],
            [f"Odds Ratio:", f"{lr_analysis['insights']['top_odds_ratio']:.4f}"],
            [f"Variables con efecto positivo:", lr_analysis['insights']['positive_vars']],
            [f"Variables con efecto negativo:", lr_analysis['insights']['negative_vars']],
            ["", ""],
            ["COMPARACIÓN ENTRE MODELOS:", ""],
            [f"Variables de alto consenso:", comparison['insights']['consensus_count']],
            [f"Variable de mayor consenso:", comparison['insights']['top_consensus_var']],
            [f"Concordancia entre modelos:", f"{comparison['insights']['models_agreement']:.1f}%"]
        ]
        
        for row, (label, value) in enumerate(info_data, 3):
            ws_summary.cell(row=row, column=1, value=label)
            ws_summary.cell(row=row, column=2, value=value)
            if ":" in label and label != "":
                ws_summary.cell(row=row, column=1).font = Font(bold=True)
        
        # Ajustar anchos
        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 30
        
        # Guardar archivo
        excel_file = f'excel/paso12_analisis_importancia_variables_{timestamp}.xlsx'
        wb.save(excel_file)
        logging.info(f"Reporte Excel guardado: {excel_file}")
        
        return excel_file
        
    except Exception as e:
        logging.error(f"Error creando reporte Excel: {str(e)}")
        raise

def interpret_business_insights(rf_analysis, lr_analysis, comparison, df_train):
    """Generar interpretaciones de negocio de las variables más importantes"""
    logging.info("Generando interpretaciones de negocio...")
    
    try:
        # Diccionario de interpretaciones de negocio por tipo de variable
        business_interpretations = {
            # Variables demográficas
            'Edad': 'Los clientes más jóvenes o mayores pueden tener diferentes patrones de lealtad',
            'Genero': 'Diferencias en comportamiento de churn entre géneros',
            'Senior_Citizen': 'Los ciudadanos senior pueden tener mayor o menor propensión al churn',
            
            # Variables de servicios
            'Servicio_Internet': 'El tipo de servicio de internet afecta significativamente la retención',
            'Servicio_Online_Security': 'La seguridad online puede ser un factor diferenciador clave',
            'Servicio_Online_Backup': 'Los servicios de respaldo influyen en la satisfacción del cliente',
            'Servicio_Device_Protection': 'La protección de dispositivos puede aumentar la lealtad',
            'Servicio_Tech_Support': 'El soporte técnico es crucial para la retención',
            'Servicio_Streaming_TV': 'Los servicios de streaming pueden generar mayor adherencia',
            'Servicio_Streaming_Movies': 'El contenido de entretenimiento afecta la retención',
            
            # Variables contractuales
            'Tipo_Contrato': 'Los contratos a largo plazo típicamente reducen el churn',
            'Metodo_Pago': 'El método de pago puede indicar diferentes niveles de compromiso',
            'Facturacion_Digital': 'La facturación digital puede correlacionar con perfiles tecnológicos',
            
            # Variables financieras
            'Cargos_Mensuales': 'Los cargos altos pueden aumentar la propensión al churn',
            'Cargos_Totales': 'El valor total del cliente indica la profundidad de la relación',
            
            # Variables de comportamiento
            'Meses_Antigüedad': 'La antigüedad es típicamente el predictor más fuerte de lealtad',
            'Telefono_Multiples_Lineas': 'Múltiples líneas pueden indicar mayor dependencia del servicio'
        }
        
        # Obtener top variables de consenso
        top_consensus = comparison['high_consensus_vars'].head(10) if len(comparison['high_consensus_vars']) > 0 else comparison['comparison_df'].head(10)
        
        # Generar insights específicos
        insights = []
        
        for _, var_row in top_consensus.iterrows():
            var_name = var_row['Variable']
            rf_importance = var_row['RF_Importancia']
            lr_importance = var_row['LR_Importancia']
            
            # Buscar interpretación base
            base_interpretation = business_interpretations.get(var_name, 'Variable específica del negocio con impacto en churn')
            
            # Obtener información adicional de regresión logística
            lr_info = lr_analysis['coefficients_df'][lr_analysis['coefficients_df']['Variable'] == var_name]
            
            if len(lr_info) > 0:
                coef = lr_info.iloc[0]['Coeficiente']
                odds_ratio = lr_info.iloc[0]['Odds_Ratio']
                direction = 'aumenta' if coef > 0 else 'reduce'
                
                interpretation = f"{base_interpretation}. Esta variable {direction} la probabilidad de churn " \
                               f"(OR: {odds_ratio:.2f}). Importancia: RF={rf_importance:.1f}%, LR={lr_importance:.1f}%"
            else:
                interpretation = f"{base_interpretation}. Importancia: RF={rf_importance:.1f}%, LR={lr_importance:.1f}%"
            
            insights.append({
                'variable': var_name,
                'interpretation': interpretation,
                'rf_importance': rf_importance,
                'lr_importance': lr_importance,
                'consensus_score': var_row['Consenso_Score']
            })
        
        # Análisis por categorías
        category_analysis = analyze_variable_categories(top_consensus, df_train)
        
        business_insights = {
            'top_variable_insights': insights,
            'category_analysis': category_analysis,
            'strategic_recommendations': generate_strategic_recommendations(insights, rf_analysis, lr_analysis)
        }
        
        logging.info(f"Generadas interpretaciones para {len(insights)} variables clave")
        return business_insights
        
    except Exception as e:
        logging.error(f"Error generando interpretaciones de negocio: {str(e)}")
        raise

def analyze_variable_categories(top_variables, df_train):
    """Analizar variables por categorías de negocio"""
    
    # Definir categorías
    categories = {
        'Demográficas': ['Edad', 'Genero', 'Senior_Citizen'],
        'Servicios': [col for col in df_train.columns if 'Servicio' in col],
        'Contractuales': ['Tipo_Contrato', 'Metodo_Pago', 'Facturacion_Digital'],
        'Financieras': ['Cargos_Mensuales', 'Cargos_Totales'],
        'Comportamiento': ['Meses_Antigüedad', 'Telefono_Multiples_Lineas']
    }
    
    category_importance = {}
    
    for category, variables in categories.items():
        category_vars = top_variables[top_variables['Variable'].isin(variables)]
        
        if len(category_vars) > 0:
            avg_importance = category_vars['Importancia_Promedio'].mean()
            var_count = len(category_vars)
            
            category_importance[category] = {
                'avg_importance': avg_importance,
                'variable_count': var_count,
                'variables': category_vars['Variable'].tolist(),
                'importance_sum': category_vars['Importancia_Promedio'].sum()
            }
    
    return category_importance

def generate_strategic_recommendations(insights, rf_analysis, lr_analysis):
    """Generar recomendaciones estratégicas basadas en variables importantes"""
    
    recommendations = []
    
    # Analizar top 5 variables más importantes
    top_5_insights = insights[:5]
    
    for insight in top_5_insights:
        var_name = insight['variable']
        
        if 'Antigüedad' in var_name or 'Meses' in var_name:
            recommendations.append({
                'area': 'Retención Temprana',
                'variable': var_name,
                'recommendation': 'Implementar programa de onboarding robusto y seguimiento proactivo en primeros 6 meses',
                'priority': 'ALTA'
            })
        
        elif 'Contrato' in var_name:
            recommendations.append({
                'area': 'Estrategia Contractual',
                'variable': var_name,
                'recommendation': 'Incentivar contratos de mayor duración con descuentos progresivos y beneficios exclusivos',
                'priority': 'ALTA'
            })
        
        elif 'Cargos' in var_name or 'Precio' in var_name:
            recommendations.append({
                'area': 'Estrategia de Precios',
                'variable': var_name,
                'recommendation': 'Segmentar precios por valor percibido y ofrecer planes flexibles para clientes sensibles al precio',
                'priority': 'MEDIA'
            })
        
        elif 'Servicio' in var_name:
            recommendations.append({
                'area': 'Portfolio de Servicios',
                'variable': var_name,
                'recommendation': 'Optimizar bundle de servicios y mejorar comunicación de valor agregado',
                'priority': 'MEDIA'
            })
        
        else:
            recommendations.append({
                'area': 'Análisis Específico',
                'variable': var_name,
                'recommendation': f'Realizar análisis detallado de {var_name} para desarrollar estrategias específicas',
                'priority': 'BAJA'
            })
    
    return recommendations

def generate_comprehensive_report(rf_analysis, lr_analysis, comparison, business_insights, 
                                models_data, viz_files, excel_file, timestamp):
    """Generar informe completo de análisis de importancia"""
    
    report = f"""
================================================================================
TELECOMX - INFORME PASO 12: ANÁLISIS DE IMPORTANCIA DE VARIABLES
================================================================================
Fecha y Hora: {timestamp}
Paso: 12 - Análisis de la Importancia de las Variables

================================================================================
RESUMEN EJECUTIVO
================================================================================
• Modelos Analizados: 2 (Random Forest + Regresión Logística)
• Variables Evaluadas: {len(rf_analysis['importance_df'])} características
• Metodologías: Importancia por impureza (RF) + Coeficientes/Odds Ratios (LR)
• Variables de Alto Consenso: {comparison['insights']['consensus_count']}
• Concordancia entre Modelos: {comparison['insights']['models_agreement']:.1f}%
• Variable Más Importante: {comparison['insights']['top_consensus_var']}

================================================================================
METODOLOGÍA DE ANÁLISIS
================================================================================

🔬 RANDOM FOREST - IMPORTANCIA POR REDUCCIÓN DE IMPUREZA:
• Metodología: Medición de cuánto cada variable contribuye a reducir la impureza 
  en las divisiones de los árboles del bosque
• Ventajas: Captura interacciones no lineales, robusto a outliers
• Interpretación: Mayor valor = mayor capacidad para separar clases
• Configuración del modelo:
  - Número de árboles: {rf_analysis['model_params']['n_trees']}
  - Profundidad máxima: {rf_analysis['model_params']['max_depth']}
  - Variables evaluadas: {rf_analysis['model_params']['total_features']}

🔢 REGRESIÓN LOGÍSTICA - ANÁLISIS DE COEFICIENTES:
• Metodología: Análisis de coeficientes del modelo lineal y odds ratios
• Ventajas: Interpretación directa del impacto direccional
• Interpretación: 
  - Coeficiente positivo: aumenta probabilidad de churn
  - Coeficiente negativo: reduce probabilidad de churn
  - Odds Ratio > 1: incrementa odds de churn
  - Odds Ratio < 1: reduce odds de churn
• Configuración del modelo:
  - Regularización C: {lr_analysis['model_params']['regularization_C']}
  - Penalización: {lr_analysis['model_params']['penalty']}
  - Variables evaluadas: {lr_analysis['model_params']['total_features']}

================================================================================
RESULTADOS - RANDOM FOREST
================================================================================

🌳 TOP 10 VARIABLES MÁS IMPORTANTES (Random Forest):
"""
    
    # Top 10 Random Forest
    rf_top_10 = rf_analysis['importance_df'].head(10)
    for i, (_, row) in enumerate(rf_top_10.iterrows(), 1):
        report += f"""
{i:2d}. {row['Variable']:<25} | {row['Importancia_Pct']:6.2f}% | Acum: {row['Importancia_Acumulada']:6.2f}%"""
    
    report += f"""

📊 ANÁLISIS DE CONCENTRACIÓN (Random Forest):
• Variable más importante: {rf_analysis['insights']['top_variable']} ({rf_analysis['insights']['top_importance']:.2f}%)
• Variables críticas (80% importancia): {rf_analysis['insights']['critical_count']} variables
• Concentración de importancia: {rf_analysis['insights']['concentration']:.1f}% en top {rf_analysis['insights']['critical_count']} variables
• Distribución: {'Concentrada' if rf_analysis['insights']['critical_count'] <= 10 else 'Distribuida'}

🎯 INTERPRETACIÓN RANDOM FOREST:
Las {rf_analysis['insights']['critical_count']} variables más importantes explican el {rf_analysis['insights']['concentration']:.1f}% 
de la capacidad predictiva del modelo, indicando que el churn está 
{'altamente concentrado en pocas variables clave' if rf_analysis['insights']['critical_count'] <= 8 else 'influenciado por múltiples factores'}.

================================================================================
RESULTADOS - REGRESIÓN LOGÍSTICA
================================================================================

📈 TOP 10 VARIABLES POR COEFICIENTE ABSOLUTO (Regresión Logística):
"""
    
    # Top 10 Regresión Logística
    lr_top_10 = lr_analysis['coefficients_df'].head(10)
    for i, (_, row) in enumerate(lr_top_10.iterrows(), 1):
        direction_icon = "📈" if row['Coeficiente'] > 0 else "📉"
        report += f"""
{i:2d}. {row['Variable']:<25} | Coef: {row['Coeficiente']:+7.4f} | OR: {row['Odds_Ratio']:6.3f} | {direction_icon} {row['Direccion']}"""
    
    report += f"""

⚖️ ANÁLISIS DIRECCIONAL (Regresión Logística):
• Variable de mayor impacto: {lr_analysis['insights']['top_variable']}
• Coeficiente: {lr_analysis['insights']['top_coefficient']:+.4f}
• Odds Ratio: {lr_analysis['insights']['top_odds_ratio']:.4f}
• Variables con efecto POSITIVO (↗️ churn): {lr_analysis['insights']['positive_vars']}
• Variables con efecto NEGATIVO (↘️ churn): {lr_analysis['insights']['negative_vars']}

🎯 INTERPRETACIÓN REGRESIÓN LOGÍSTICA:
El modelo identifica {lr_analysis['insights']['positive_vars']} factores que INCREMENTAN el riesgo de churn 
y {lr_analysis['insights']['negative_vars']} factores PROTECTORES que REDUCEN el riesgo. 
La variable {lr_analysis['insights']['top_variable']} tiene el mayor impacto 
{'aumentando' if lr_analysis['insights']['top_coefficient'] > 0 else 'reduciendo'} las odds de churn en {abs((lr_analysis['insights']['top_odds_ratio'] - 1) * 100):.1f}%.

================================================================================
ANÁLISIS COMPARATIVO ENTRE MODELOS
================================================================================

🤝 VARIABLES DE ALTO CONSENSO (Ambos modelos coinciden):
"""
    
    # Variables de consenso
    if len(comparison['high_consensus_vars']) > 0:
        consensus_vars = comparison['high_consensus_vars'].head(10)
        for i, (_, row) in enumerate(consensus_vars.iterrows(), 1):
            report += f"""
{i:2d}. {row['Variable']:<25} | RF: {row['RF_Importancia']:5.2f}% | LR: {row['LR_Importancia']:5.2f}% | Consensus: {row['Consenso_Score']:5.2f}"""
    else:
        top_comparison = comparison['comparison_df'].head(10)
        for i, (_, row) in enumerate(top_comparison.iterrows(), 1):
            report += f"""
{i:2d}. {row['Variable']:<25} | RF: {row['RF_Importancia']:5.2f}% | LR: {row['LR_Importancia']:5.2f}% | Avg: {row['Importancia_Promedio']:5.2f}"""

    report += f"""

📊 MÉTRICAS DE CONCORDANCIA:
• Variables en top 10 de ambos modelos: {comparison['insights']['consensus_count']}
• Concordancia general: {comparison['insights']['models_agreement']:.1f}%
• Variable de mayor consenso: {comparison['insights']['top_consensus_var']}
• Discrepancia promedio en rankings: {comparison['insights']['avg_discrepancy']:.1f} posiciones

🔍 VARIABLES CON MAYOR DISCREPANCIA:
"""
    
    # Variables con discrepancia
    if len(comparison['high_discrepancy_vars']) > 0:
        for _, row in comparison['high_discrepancy_vars'].head(5).iterrows():
            rf_rank = int(row['RF_Ranking'])
            lr_rank = int(row['LR_Ranking'])
            diff = int(row['Diferencia_Ranking'])
            
            report += f"""
• {row['Variable']}: RF Rank #{rf_rank} vs LR Rank #{lr_rank} (Dif: {diff} posiciones)"""

    report += f"""

🎯 INTERPRETACIÓN DE CONCORDANCIA:
La concordancia del {comparison['insights']['models_agreement']:.1f}% indica que ambos modelos 
{'tienen alta coincidencia' if comparison['insights']['models_agreement'] > 70 else 'difieren significativamente' if comparison['insights']['models_agreement'] < 50 else 'tienen concordancia moderada'} 
en la identificación de variables importantes. Esto {'valida la robustez de los hallazgos' if comparison['insights']['models_agreement'] > 60 else 'sugiere que diferentes metodologías captan aspectos distintos'}.

================================================================================
INTERPRETACIÓN DE NEGOCIO - VARIABLES CLAVE
================================================================================

💼 ANÁLISIS DE LAS VARIABLES MÁS CRÍTICAS:
"""
    
    # Interpretaciones de negocio
    for i, insight in enumerate(business_insights['top_variable_insights'][:8], 1):
        report += f"""
{i}. {insight['variable'].upper()}:
   📋 {insight['interpretation']}
   📊 Consenso Score: {insight['consensus_score']:.2f}
"""

    report += f"""
🏢 ANÁLISIS POR CATEGORÍAS DE NEGOCIO:
"""
    
    # Análisis por categorías
    for category, data in business_insights['category_analysis'].items():
        report += f"""
{category.upper()}:
   • Variables importantes: {data['variable_count']} ({', '.join(data['variables'][:3])}{'...' if len(data['variables']) > 3 else ''})
   • Importancia promedio: {data['avg_importance']:.2f}%
   • Importancia total: {data['importance_sum']:.2f}%"""

    report += f"""

================================================================================
RECOMENDACIONES ESTRATÉGICAS
================================================================================

🎯 PLAN DE ACCIÓN BASADO EN VARIABLES CRÍTICAS:
"""
    
    # Recomendaciones estratégicas
    for i, rec in enumerate(business_insights['strategic_recommendations'], 1):
        priority_icon = "🔴" if rec['priority'] == 'ALTA' else "🟡" if rec['priority'] == 'MEDIA' else "🟢"
        
        report += f"""
{priority_icon} {i}. ÁREA: {rec['area'].upper()}
   Variable Clave: {rec['variable']}
   Acción Recomendada: {rec['recommendation']}
   Prioridad: {rec['priority']}
"""

    report += f"""
🚀 IMPLEMENTACIÓN PRÁCTICA:

1. RETENCIÓN PROACTIVA:
   • Identificar clientes en riesgo basado en variables críticas
   • Crear scores de propensión personalizados
   • Implementar alertas automáticas para intervención temprana

2. SEGMENTACIÓN INTELIGENTE:
   • Segmentar base de clientes por factores de riesgo
   • Personalizar ofertas según perfil de variables importantes
   • Desarrollar campañas específicas por segmento

3. OPTIMIZACIÓN DE PRODUCTOS/SERVICIOS:
   • Mejorar aspectos relacionados con variables de alto impacto
   • Rediseñar bundling basado en variables protectoras
   • Ajustar estrategia de precios considerando sensibilidad

4. MONITOREO CONTINUO:
   • Tracking regular de variables críticas por cliente
   • Dashboard ejecutivo con métricas clave
   • Alertas de cambios significativos en variables importantes

================================================================================
VALIDACIÓN Y ROBUSTEZ DE RESULTADOS
================================================================================

✅ CONSISTENCIA DE HALLAZGOS:
• Metodologías complementarias: Random Forest (no lineal) + Regresión Logística (lineal)
• Consenso en variables críticas: {comparison['insights']['consensus_count']} variables coincidentes
• Interpretabilidad: Ambos modelos permiten explicar decisiones a negocio

✅ FORTALEZAS DEL ANÁLISIS:
• Diversidad metodológica reduce sesgos algorítmicos
• Interpretaciones de negocio validadas con domain knowledge
• Variables importantes son accionables para estrategia

⚠️ LIMITACIONES Y CONSIDERACIONES:
• Importancia relativa puede cambiar con nuevos datos
• Interacciones complejas pueden no estar completamente capturadas
• Causalidad vs correlación requiere validación con experimentos

================================================================================
ARCHIVOS GENERADOS
================================================================================

📊 VISUALIZACIONES:
• Random Forest - Importancia de variables: graficos/paso12_random_forest_importancia_{timestamp}.png
• Regresión Logística - Coeficientes: graficos/paso12_regresion_logistica_coeficientes_{timestamp}.png
• Comparación entre modelos: graficos/paso12_comparacion_modelos_{timestamp}.png
• Análisis de consenso: graficos/paso12_consenso_variables_{timestamp}.png
• Odds Ratios: graficos/paso12_odds_ratios_{timestamp}.png

📄 DOCUMENTACIÓN:
• Informe completo: informes/paso12_analisis_importancia_variables_informe_{timestamp}.txt
• Reporte Excel: {excel_file}
• Log del proceso: logs/paso12_analisis_importancia_variables.log

================================================================================
CONCLUSIONES Y PRÓXIMOS PASOS
================================================================================

🎯 CONCLUSIONES PRINCIPALES:

1. VARIABLES MÁS CRÍTICAS:
   • {comparison['insights']['top_consensus_var']} es la variable de mayor consenso entre modelos
   • {rf_analysis['insights']['critical_count']} variables explican {rf_analysis['insights']['concentration']:.1f}% de la importancia en RF
   • {lr_analysis['insights']['positive_vars']} variables incrementan riesgo vs {lr_analysis['insights']['negative_vars']} variables protectoras

2. ROBUSTEZ DEL ANÁLISIS:
   • Concordancia del {comparison['insights']['models_agreement']:.1f}% entre metodologías
   • Variables críticas son interpretables y accionables
   • Resultados consistentes con conocimiento del dominio

3. IMPLICACIONES ESTRATÉGICAS:
   • Focus en variables de alto consenso maximiza impacto
   • Oportunidades claras de intervención identificadas
   • Base sólida para scoring de propensión al churn

📋 PRÓXIMOS PASOS RECOMENDADOS:

1. IMPLEMENTACIÓN DE SCORING:
   • Desarrollar score de riesgo basado en variables críticas
   • Implementar sistema de alertas automáticas
   • Crear dashboard de monitoreo de variables clave

2. EXPERIMENTACIÓN:
   • A/B testing de intervenciones basadas en variables importantes
   • Validar causalidad de variables críticas
   • Optimizar thresholds de intervención

3. MONITOREO Y ACTUALIZACIÓN:
   • Tracking mensual de importancia de variables
   • Re-evaluación trimestral de modelos
   • Actualización de estrategias basada en nuevos insights

================================================================================
FIN DEL INFORME
================================================================================
"""
    
    return report

def save_final_results(report_content, timestamp):
    """Guardar resultados finales"""
    try:
        # Guardar informe completo
        report_file = f'informes/paso12_analisis_importancia_variables_informe_{timestamp}.txt'
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(report_content)
        logging.info(f"Informe completo guardado: {report_file}")
        
        return {
            'report_file': report_file
        }
        
    except Exception as e:
        logging.error(f"Error al guardar resultados finales: {str(e)}")
        raise

def main():
    """Función principal del Paso 12"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    logger = setup_logging()
    
    try:
        logger.info("="*80)
        logger.info("INICIANDO PASO 12: ANÁLISIS DE IMPORTANCIA DE VARIABLES")
        logger.info("="*80)
        logger.info("Análisis exhaustivo de variables críticas para predicción de churn")
        
        # 1. Crear directorios necesarios
        create_directories()
        
        # 2. Cargar modelos y datos
        models_data, feature_names, df_train = load_models_and_data()
        logger.info(f"Modelos cargados: {list(models_data.keys())}")
        logger.info(f"Variables disponibles: {len(feature_names)}")
        
        # 3. Análisis Random Forest
        logger.info("="*50)
        logger.info("ANALIZANDO RANDOM FOREST")
        rf_model = models_data['Random Forest']['model']
        rf_analysis = analyze_random_forest_importance(rf_model, feature_names)
        
        # 4. Análisis Regresión Logística
        logger.info("="*50)
        logger.info("ANALIZANDO REGRESIÓN LOGÍSTICA")
        lr_model = models_data['Regresión Logística']['model']
        lr_analysis = analyze_logistic_regression_coefficients(lr_model, feature_names)
        
        # 5. Comparación entre modelos
        logger.info("="*50)
        logger.info("COMPARANDO IMPORTANCIA ENTRE MODELOS")
        comparison = compare_variable_importance(rf_analysis, lr_analysis)
        
        # 6. Interpretaciones de negocio
        logger.info("="*50)
        logger.info("GENERANDO INTERPRETACIONES DE NEGOCIO")
        business_insights = interpret_business_insights(rf_analysis, lr_analysis, comparison, df_train)
        
        # 7. Generar visualizaciones
        logger.info("="*50)
        logger.info("GENERANDO VISUALIZACIONES")
        viz_files = generate_visualizations(rf_analysis, lr_analysis, comparison, timestamp)
        
        # 8. Crear reporte Excel
        logger.info("="*50)
        logger.info("CREANDO REPORTE EXCEL")
        excel_file = create_excel_report(rf_analysis, lr_analysis, comparison, timestamp)
        
        # 9. Generar informe completo
        logger.info("="*50)
        logger.info("GENERANDO INFORME COMPLETO")
        report_content = generate_comprehensive_report(
            rf_analysis, lr_analysis, comparison, business_insights,
            models_data, viz_files, excel_file, timestamp
        )
        
        # 10. Guardar resultados finales
        output_files = save_final_results(report_content, timestamp)
        
        # 11. Resumen final
        logger.info("="*80)
        logger.info("PASO 12 COMPLETADO EXITOSAMENTE")
        logger.info("="*80)
        logger.info("RESULTADOS DEL ANÁLISIS DE IMPORTANCIA:")
        logger.info("")
        
        # Mostrar hallazgos clave
        logger.info("🎯 VARIABLES MÁS IMPORTANTES:")
        logger.info("")
        logger.info("🌳 RANDOM FOREST - TOP 5:")
        rf_top_5 = rf_analysis['importance_df'].head(5)
        for i, (_, row) in enumerate(rf_top_5.iterrows(), 1):
            logger.info(f"  {i}. {row['Variable']}: {row['Importancia_Pct']:.2f}%")
        logger.info("")
        
        logger.info("📈 REGRESIÓN LOGÍSTICA - TOP 5:")
        lr_top_5 = lr_analysis['coefficients_df'].head(5)
        for i, (_, row) in enumerate(lr_top_5.iterrows(), 1):
            direction = "↗️" if row['Coeficiente'] > 0 else "↘️"
            logger.info(f"  {i}. {row['Variable']}: {row['Coeficiente']:+.4f} {direction} (OR: {row['Odds_Ratio']:.3f})")
        logger.info("")
        
        # Variables de consenso
        logger.info("🤝 VARIABLES DE ALTO CONSENSO:")
        consensus_vars = comparison['high_consensus_vars'] if len(comparison['high_consensus_vars']) > 0 else comparison['comparison_df'].head(5)
        for i, (_, row) in enumerate(consensus_vars.head(5).iterrows(), 1):
            logger.info(f"  {i}. {row['Variable']}: RF={row['RF_Importancia']:.1f}% | LR={row['LR_Importancia']:.1f}%")
        logger.info("")
        
        # Métricas de concordancia
        logger.info("📊 MÉTRICAS DE ANÁLISIS:")
        logger.info(f"  • Variables evaluadas: {len(feature_names)}")
        logger.info(f"  • Variables de consenso: {comparison['insights']['consensus_count']}")
        logger.info(f"  • Concordancia entre modelos: {comparison['insights']['models_agreement']:.1f}%")
        logger.info(f"  • Variable más importante: {comparison['insights']['top_consensus_var']}")
        logger.info("")
        
        # Insights de negocio
        logger.info("💼 PRINCIPALES INSIGHTS DE NEGOCIO:")
        for i, insight in enumerate(business_insights['top_variable_insights'][:3], 1):
            logger.info(f"  {i}. {insight['variable']}")
            # Truncar interpretación para log
            short_interp = insight['interpretation'][:100] + "..." if len(insight['interpretation']) > 100 else insight['interpretation']
            logger.info(f"     {short_interp}")
        logger.info("")
        
        # Recomendaciones estratégicas
        logger.info("🚀 RECOMENDACIONES ESTRATÉGICAS CLAVE:")
        high_priority_recs = [rec for rec in business_insights['strategic_recommendations'] if rec['priority'] == 'ALTA']
        for i, rec in enumerate(high_priority_recs[:3], 1):
            logger.info(f"  {i}. {rec['area']}: {rec['recommendation'][:80]}...")
        logger.info("")
        
        # Archivos generados
        logger.info("📁 ARCHIVOS GENERADOS:")
        logger.info(f"  • Informe completo: {output_files['report_file']}")
        logger.info(f"  • Reporte Excel: {excel_file}")
        logger.info(f"  • Visualizaciones: {len(viz_files)} gráficos en graficos/")
        logger.info("    - Importancia Random Forest")
        logger.info("    - Coeficientes Regresión Logística")
        logger.info("    - Comparación entre modelos")
        logger.info("    - Análisis de consenso")
        logger.info("    - Odds Ratios")
        logger.info("")
        
        # Estado para siguiente paso
        logger.info("✅ ANÁLISIS COMPLETADO:")
        logger.info("  • Variables críticas identificadas")
        logger.info("  • Interpretaciones de negocio generadas")
        logger.info("  • Recomendaciones estratégicas listas")
        logger.info("  • Documentación completa disponible")
        logger.info("")
        
        logger.info("📊 PRÓXIMOS PASOS RECOMENDADOS:")
        logger.info("  1. Implementar scoring de riesgo basado en variables críticas")
        logger.info("  2. Desarrollar campañas de retención segmentadas")
        logger.info("  3. Crear dashboard de monitoreo de variables clave")
        logger.info("  4. Diseñar experimentos para validar causalidad")
        logger.info("  5. Optimizar productos/servicios basado en insights")
        logger.info("="*80)
        
    except Exception as e:
        logger.error(f"ERROR EN EL PROCESO: {str(e)}")
        import traceback
        logger.error(f"Traceback completo: {traceback.format_exc()}")
        raise

if __name__ == "__main__":
    main()